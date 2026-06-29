from datetime import datetime
import os
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent

EXPORT_DIR = Path(os.environ.get("PROGETTO_CAVE_EXPORT_DIR", BASE_DIR / "export"))
EXPORT_PATH = EXPORT_DIR / "report.xlsx"


def _pulisci_valore_excel(valore):
    if isinstance(valore, str):
        return ILLEGAL_CHARACTERS_RE.sub("", valore)
    return valore


def _pulisci_dataframe_excel(df):
    df_pulito = df.copy()
    for colonna in df_pulito.select_dtypes(include=["object", "string"]).columns:
        df_pulito[colonna] = df_pulito[colonna].map(_pulisci_valore_excel)
    return df_pulito


def _normalizza_nome_colonna(nome):
    return str(nome).replace("_", " ").title()


def _scrivi_riga_su_foglio(ws, valori):
    valori = list(valori)
    ws.append(valori)
    indice_riga = ws._current_row

    for indice_colonna, valore in enumerate(valori, start=1):
        if isinstance(valore, str) and valore.startswith("="):
            ws.cell(indice_riga, indice_colonna).data_type = "s"


def _stile_intestazione(ws):
    header_fill = PatternFill("solid", fgColor="173F35")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D7DEDA"))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 24


def _formatta_foglio_principale(ws, config_report):
    ws.title = config_report.get("main_sheet", "Report")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = config_report.get("freeze_panes", "A2")

    original_headers = {cell.value: cell.column for cell in ws[1]}
    larghezze = config_report.get("column_widths", {})
    wrap_columns = set(config_report.get("wrap_columns", []))
    number_formats = config_report.get("number_formats", {})
    soft_fill = PatternFill("solid", fgColor="F6F8F7")
    thin_border = Border(bottom=Side(style="thin", color="D7DEDA"))

    for cell in ws[1]:
        nome_originale = cell.value
        ws.column_dimensions[cell.column_letter].width = larghezze.get(nome_originale, 18)
        cell.value = _normalizza_nome_colonna(nome_originale)

    _stile_intestazione(ws)
    if config_report.get("fast_format"):
        for colonna, formato in number_formats.items():
            col_index = original_headers.get(colonna)
            if not col_index:
                continue
            for row in ws.iter_rows(
                min_row=2,
                min_col=col_index,
                max_col=col_index,
            ):
                row[0].number_format = formato
        return

    for row in ws.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = soft_fill

        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        for colonna in wrap_columns:
            col_index = original_headers.get(colonna)
            if col_index:
                row[col_index - 1].alignment = Alignment(wrap_text=True, vertical="center")

        for colonna, formato in number_formats.items():
            col_index = original_headers.get(colonna)
            if col_index:
                row[col_index - 1].number_format = formato

    val_tot_col = original_headers.get("Val_tot")
    if val_tot_col and ws.max_row > 1:
        lettera = get_column_letter(val_tot_col)
        ws.conditional_formatting.add(
            f"{lettera}2:{lettera}{ws.max_row}",
            CellIsRule(operator="equal", formula=["0"], font=Font(color="7A817D")),
        )


def _scrivi_riepilogo(wb, df, summary_config, index):
    sheet_name = summary_config.get("sheet_name", f"Riepilogo_{index}")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    group_by = summary_config.get("group_by", [])
    aggregazioni = summary_config.get("aggregations", {})
    if not group_by or not aggregazioni:
        return

    colonne = group_by + list(aggregazioni)
    riepilogo = (
        df.groupby(group_by, dropna=False, as_index=False)
        .agg(aggregazioni)
        .sort_values(group_by, kind="stable")
    )

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    _scrivi_riga_su_foglio(
        ws,
        [_normalizza_nome_colonna(colonna) for colonna in colonne],
    )

    for _, row in riepilogo[colonne].iterrows():
        _scrivi_riga_su_foglio(
            ws,
            [_pulisci_valore_excel(valore) for valore in row.tolist()],
        )

    _stile_intestazione(ws)
    thin_border = Border(bottom=Side(style="thin", color="D7DEDA"))
    number_formats = summary_config.get("number_formats", {})

    for cell in ws[1]:
        ws.column_dimensions[cell.column_letter].width = 18

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        for posizione, colonna in enumerate(colonne):
            formato = number_formats.get(colonna)
            if formato:
                row[posizione].number_format = formato


def _scrivi_dataframe_su_foglio(ws, df):
    for row in dataframe_to_rows(df, index=False, header=True):
        _scrivi_riga_su_foglio(ws, row)


def _salva_excel(percorso, df, config_report):
    wb = Workbook()
    ws = wb.active
    _scrivi_dataframe_su_foglio(ws, df)
    _formatta_foglio_principale(ws, config_report)

    for extra_sheet in config_report.get("extra_sheets", []):
        _scrivi_foglio_dataframe(wb, extra_sheet["df"], extra_sheet["config"])

    for index, summary_config in enumerate(config_report.get("summaries", []), start=1):
        _scrivi_riepilogo(wb, df, summary_config, index)

    wb.active = wb.sheetnames.index(config_report.get("main_sheet", "Report"))
    wb.save(percorso)


def _scrivi_foglio_dataframe(wb, df, sheet_config):
    sheet_name = sheet_config.get("main_sheet", sheet_config.get("name", "Foglio"))
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    df = _pulisci_dataframe_excel(df)
    _scrivi_dataframe_su_foglio(ws, df)

    _formatta_foglio_principale(ws, sheet_config)


def _percorso_export(config_report):
    nome_file = config_report.get("output_file", EXPORT_PATH.name) if config_report else EXPORT_PATH.name
    nome_file = nome_file.format(**config_report)
    return EXPORT_DIR / nome_file


def esegui(df, config_report=None):
    config_report = config_report or {}
    df = _pulisci_dataframe_excel(df)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    percorso_base = _percorso_export(config_report)

    try:
        percorso = percorso_base
        _salva_excel(percorso, df, config_report)
    except PermissionError:
        percorso = percorso_base.with_name(
            f"{percorso_base.stem}_{datetime.now():%Y%m%d_%H%M%S}{percorso_base.suffix}"
        )
        _salva_excel(percorso, df, config_report)
        print(f"Report principale bloccato. Creata una copia: {percorso}")
        return percorso

    print(f"Conversione completata. File creato: {percorso}")
    return percorso
